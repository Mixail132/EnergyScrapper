# This is a common script where all the logic connected together
# This is for making exe file
# This works for the period

import os
import fdb
from dotenv import load_dotenv


# ----------------------------connections.py--------------------------
load_dotenv()
DATABASE = os.getenv("DATABASE_PATH")
DB_CLIENT = os.getenv("DBCLIENT_PATH")


connect = fdb.connect(
    host='localhost',
    database=DATABASE,
    user='sysdba',
    password='masterkey',
    charset="UTF-8",
    fb_library_name=DB_CLIENT
)

cursor = connect.cursor()

#----------------------------counters.py-----------------------------
import re
# from connections import cursor


def get_counters_info():
    """ Gets counters' info (number of a counter, what it counts)
        from the database, filters the extra data and makes a list
        of the data. """
    request = ("SELECT NAME, SERIALNUM, NUM_DEVICE, NUM_CH FROM CHANNEL")
    cursor.execute(request)
    all_counters_info = []
    for i in cursor.fetchall():
        counters_info = []
        for j in i:
            counters_info.append((j))
        reg = r"R\-|R\+|A\-|А\-"
        if not re.findall(reg, str(counters_info[0])):
            all_counters_info.append(counters_info)
    return all_counters_info

#----------------------------consunptions.py---------------------------------
from datetime import datetime, timedelta
# from connections import cursor


def get_counters_consumption(data_date):
    """ Gets all the energy consumption data from the database. """
    request = (f"""
        SELECT 
        NUM_DEVICE,
        NUM_CH, 
        DT_DAY,
        TARIF1, 
        TARIF2, 
        TARIF3 
        FROM IMPULS4 
        WHERE 
        DT_DAY='{data_date}'
        """)
    cursor.execute(request)
    return cursor.fetchall()

#----------------------------excels.py-------------------------------
from openpyxl import load_workbook
# import os
# import datetime
# from dotenv import load_dotenv
# from views import retreived_energy_handler


# load_dotenv()


def put_consumptions_to_excel(data_day, counters_data):
    counters_storage_file = os.getenv('DBEXCEL_PATH')
    counters_data_storage =  load_workbook(counters_storage_file) 
    # today_date = datetime.datetime.now()
    current_date_row = data_day.toordinal()-737328
    counters_numbers={}
    for col_number in range(2, 47):
        counter_number = counters_data_storage.active.cell(row=3, column = col_number).value
        counters_numbers[counter_number] = col_number
    for counter_number, col_number in counters_numbers.items():
        try:
            energy_value = (counters_data[str(counter_number)])
        except KeyError:
            continue
        column_number = counters_numbers[counter_number]  
        counters_data_storage.active.cell(row=current_date_row, column=column_number).value=energy_value       
    counters_data_storage.save(counters_storage_file)

#--------------------------------views.py-------------------------------------
# from datetime import datetime, timedelta
# from consumptions import get_counters_consumption
# from counters import get_counters_info
# from excels import put_consumptions_to_excel


def day_energy_handler(sql_date, data_day):
    """ Gets energy consumption data from the database
        looks for which the data belongs to which counter,
        makes a dict of the summed up data,
        puts the data to excel file. """
    consumptions = get_counters_consumption(sql_date)
    counters = get_counters_info()
    filtered_consumption={} 
    for counter in counters:
        for consumption in consumptions:
            if counter[3] == consumption[1] and counter[2] == consumption[0]:
                one_counter_total_energy = (sum(consumption[3:6]))
                filtered_consumption[counter[1]] = round(one_counter_total_energy, 2)
                break
    put_consumptions_to_excel(data_day, filtered_consumption)
    return filtered_consumption


def period_energy_handler():
    """ Counts the period of needed data,
        pass every single data as a parameter
        to a single day retrieving energy function. """
    number_of_days=input("Enter nuber of days for updating")
    for day in range(int(number_of_days), -1, -1): # Today to three days ago in descendштп order
        data_day = datetime.today()-timedelta(days=day)
        sql_date = data_day.strftime("%Y-%m-%d")
        day_energy_handler(sql_date, data_day)

#----------------------------------START-------------------------------------
try:
    period_energy_handler()
except PermissionError:
    print(f"Excel файл {os.getenv('DBEXCEL_PATH')} открыт. \n Закройте его!" )
