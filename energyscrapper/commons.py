# This file contains all the logic connected together for making an exe file later

# ------------------------------------- .env --------------------------------------------
# DATABASE_PATH = "C://Program Files (x86)/Control Center Server/Base/ENERGY.GDB"
# DBCLIENT_PATH = "C://Program Files (x86)/Firebird/Firebird_2_1/bin/fbclient.dll"
# DBEXCEL_PATH = "C://Users/Ev/Documents/Python/EnergyScrapper/data/energy.xlsx"
DATABASE_PATH = 'C:\\Program Files (x86)\\Control Center Server\\Base\\ENERGY.GDB'
DBCLIENT_PATH = 'C:\\Program Files (x86)\\Firebird\\Firebird_2_1\\bin\\fbclient.dll'
DBEXCEL_PATH = 'C:\\Users\\User\\Documents\\Информация\\Расчеты\\Энергопотребление\\Показания\\Разное\\energy.xlsx'
# ------------------------------------ dirs.py ------------------------------------------
from pathlib import Path

DATABASE = str(Path(DATABASE_PATH))
DB_CLIENT = str(Path(DBCLIENT_PATH))
DB_EXCEL = str(Path(DBEXCEL_PATH))


# ---------------------------------- connections.py -------------------------------------
import fdb

connect = fdb.connect(
    host='localhost',
    database=DATABASE,
    user='sysdba',
    password='masterkey',
    charset="UTF-8",
    fb_library_name=DB_CLIENT
)

cursor = connect.cursor()

# ------------------------------- consumptions.py -----------------------------------------
def get_counters_consumption(data_date: str) -> cursor:
    today_date = datetime.today().strftime("%Y-%m-%d")
    request = (f"""
        SELECT 
        NUM_DEVICE,
        NUM_CH, 
        DT_DAY,
        TARIF1, 
        TARIF2, 
        TARIF3 
        FROM IMPULS4 
        WHERE 
        DT_DAY>='{data_date}'
        AND
        DT_DAY<='{today_date}'
        ORDER BY DT_DAY 
        """)
    cursor.execute(request)
    return cursor.fetchall()


def filter_counters_consumption(counters: list, consumptions: list) -> list:
    all_days_filtered_consumption = []
    for consumption in consumptions:
        one_day_filtered_consumption = []
        for counter in counters:
            if counter[3] == consumption[1] and counter[2] == consumption[0]:
                one_day_filtered_consumption.append(consumption[2])
                one_day_filtered_consumption.append(counter[1])
                one_counter_total_energy = (sum(consumption[3:6]))
                one_day_filtered_consumption.append(round(one_counter_total_energy, 2))
                break
        if one_day_filtered_consumption:
            all_days_filtered_consumption.append(one_day_filtered_consumption)
    return all_days_filtered_consumption


def make_consumptions_per_date(period_date: str, consumptions: list) -> dict:
    date = datetime.strptime(period_date, "%Y-%m-%d")
    days = (datetime.today() - date).days
    all_days = {}
    for day in reversed(range(days)):
        get_day = date.today() - timedelta(days=day)
        one_day_consumption = {}
        for consumption in consumptions:
            if consumption[0].date() == get_day.date():
                one_day_consumption[consumption[1]] = consumption[2]
        if one_day_consumption:
            all_days[f'{get_day.date()}'] = one_day_consumption
    return all_days

# -------------------------- counters.py ------------------------------------
import re


def get_counters_info():
    request = ("""
            SELECT 
            NAME, 
            SERIALNUM, 
            NUM_DEVICE, 
            NUM_CH 
            FROM CHANNEL
            """)
    cursor.execute(request)
    all_counters_info = []
    for i in cursor.fetchall():
        counters_info = []
        for j in i:
            counters_info.append(j)
        reg = r"R\-|R\+|A\-|А\-"
        if not re.findall(reg, str(counters_info[0])):
            all_counters_info.append(counters_info)
    return all_counters_info

# ------------------------------ excels.py -------------------------------------
from openpyxl import load_workbook


def put_consumptions_to_excel(counters_data: dict) -> None:
    counters_storage_file = DB_EXCEL
    counters_data_storage = load_workbook(counters_storage_file)
    for date, counters_values in counters_data.items():
        date_to_paste = datetime.strptime(date, "%Y-%m-%d")
        current_date_row = date_to_paste.toordinal()-737328
        counters_numbers = {}
        for col_number in range(2, 47):
            counter_number = counters_data_storage.active.cell(row=3, column=col_number).value
            counters_numbers[counter_number] = col_number
        for counter_number, col_number in counters_numbers.items():
            try:
                energy_value = (counters_values[str(counter_number)])
            except KeyError:
                continue
            column_number = counters_numbers[counter_number]
            counters_data_storage.active.cell(row=current_date_row, column=column_number).value = energy_value
    counters_data_storage.save(counters_storage_file)


# ------------------------------- views.py ----------------------------------
from datetime import datetime, timedelta

start_day = datetime.today()
# start_day = datetime(2024, 2, 13)


def period_energy_handler() -> dict:
    data_day = start_day-timedelta(days=33)
    sql_date = data_day.strftime("%Y-%m-%d")
    all_consumption = get_counters_consumption(sql_date)
    counters_info = get_counters_info()
    filtered_consumption = filter_counters_consumption(counters_info, all_consumption)
    needed_consumption = make_consumptions_per_date(sql_date, filtered_consumption)
    put_consumptions_to_excel(needed_consumption)
    return needed_consumption


period_energy_handler()
