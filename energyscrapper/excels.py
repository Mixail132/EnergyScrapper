# Data storage in excel handler
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv
import dirs


load_dotenv()


def put_consumptions_to_excel(counters_data: dict) -> None:
    """
    Gets the dict of the data.
    Opens the Excel file and finds what counter defers to what cell.
    Finds the current data row.
    Inserts the appropriate data into the appropriate rows.
    Saves the Excel file.
    :param counters_data: looks like this:
    2024-02-04 {'18085423': 14239.49, '19094836': 20873.95, ... '18086745': 9891.18}
    2024-02-05 {'18085423': 14255.95, '19094836': 20892.65, ... '18086745': 9906.58}
    where:
        2024-02-04 - the meters' reading date;
        18085423   - one of the counter factory number;
        14239.49   - the meter's value for the date.
    """
    counters_storage_file = dirs.DB_EXCEL
    counters_data_storage = load_workbook(counters_storage_file)
    for date, counters_values in counters_data.items():
        date_to_paste = datetime.strptime(date, "%Y-%m-%d")
        current_date_row = date_to_paste.toordinal()-737328
        counters_numbers = {}
        for col_number in range(2, 47):
            counter_number = counters_data_storage.active.cell(row=3, column=col_number).value
            counters_numbers[counter_number] = col_number
        for counter_number, col_number in counters_numbers.items():
            try:
                energy_value = (counters_values[str(counter_number)])
            except KeyError:
                continue
            column_number = counters_numbers[counter_number]
            counters_data_storage.active.cell(row=current_date_row, column=column_number).value = energy_value
    counters_data_storage.save(counters_storage_file)
