import requests
import datetime
import time
from openpyxl import load_workbook

now = datetime.datetime.now()
bef = now - datetime.timedelta(days=7)
stnow = now.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3]
stbef = bef.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3]

# убрать бесконечное получение session id при каждом запуске скрипта
# сделать словарь с остальными необходимыми данными

url_2 = "http://37.17.58.181"
url_3 = f"http://37.17.58.181/Report/ViewReport?ReportIndex=ictweb5.Domain.Reports.Heat.LocationHeatCurrentDataReportSQLDataRepositoryClass&Regions=&Locations=&Objects=&History=0&ArchiveType=1&AccountingType=0&DateFrom={stbef}Z&DateTo={stnow}Z"
url_4 = "http://37.17.58.181/Schedule/ReadArchive?Regions=&Locations=&Objects=&Type=1&DayCount=7&HourCount=0&MinuteCount=0&EventCount=undefined&ReadUnSuccessful=undefined"
headers_1 = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br', 'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8', 'Connection': 'keep-alive',
    'Host': 'indel.becloud.by', 'sec-ch-ua': '"Not_A Brand";v="99", "Microsoft Edge";v="109", "Chromium";v="109"',
    'sec-ch-ua-mobile': '?0', 'sec-ch-ua-platform': '"Windows"', 'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate', 'Sec-Fetch-Site': 'none', 'Sec-Fetch-User': '?1', 'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.115', }

data = {"Name": "арена", "Password": "885522"}

headers_3 = {'Accept': '*/*', 'Accept-Encoding': 'gzip, deflate, br', 'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8',
    'Connection': 'keep-alive', 'Content-Length': '0', 'Host': '37.17.58.181', 'Origin': 'http://37.17.58.181',
    'Referer': 'http://37.17.58.181/Report/Index',
    'sec-ch-ua': '"Not_A Brand";v="99", "Microsoft Edge";v="109", "Chromium";v="109"', 'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"', 'Sec-Fetch-Dest': 'empty', 'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.115',
    'X-Requested-With': 'XMLHttpRequest'}

# отправляем первый запрос (на получение cookie)
set = requests.get(url_2, headers=headers_1)

# получаем в ответ cookie (ID сессии)
set_cookie = set.headers["Set-Cookie"]

# выделяем из полученных cookie нужный фрагмент
set_cookies = set_cookie[:set_cookie.index(";")]

# добавляем cookie к заголовку следующего запроса
headers_3["Cookie"] = set_cookies

# отправляем второй запрос (на вход в систему)
ent = requests.post(url_2, headers=headers_3, data=data)

# изменяем cookie в заголовке следующго запроса
headers_3["Cookie"] = set_cookies

# отправляем следующий запрос (на опрос данных из INDEL за период)
qst = requests.post(url_4, headers=headers_3)
# time.sleep(600)

# отправляем следующий запрос (на изъятие данных)
ret = requests.post(url_3, headers=headers_3, )  # data = data)

# создаем объект (ответ в текстом виде)
res = ret.text

# отбрасываем лишнее (нужные данные внутри таблице)
bod = res[res.index("<tbody") + 26: res.index("/tbody>") - 1]

# удаляем ненужные, заменяем на "," нужные символы и превращаем итоговую строку в список
lis = bod.replace("<th>", "").replace("<tr>", "").replace("</tr>", "").replace("</th>", "").replace("<td>", "").replace(
    "</td>", ",").split(",")

# отправляем запрос на выход из системы
qut = requests.get(url="http://37.17.58.181/Auth/Logoff", headers=headers_3)

# работаем с архивом в формате excel
path = "C:/Users/User/Documents/Организации/Подрядчики/Наладка/Диспетчеризация/Индел/indel/indel.xlsx"
indel = load_workbook(path)

# находим номер строки для вставки значений в соответствии с текущей датой
dat = now.toordinal() - 737553

# создаем новый список для частично оцифрованных значений
sis = []

# находим значения с одной точкой и меняем их на число
for tex in lis:
    if "." in tex:
        if tex.count(".") == 1:
            sis.append(float(tex))
        else:
            sis.append(tex)
    else:
        sis.append(tex)

# выгружаем значения в архив
for col in range(1, 14):
    indel.active.cell(row=dat, column=col).value = sis[col + 5]
for col in range(14, 27):
    indel.active.cell(row=dat, column=col).value = sis[col + 11]

# сохраняем архив
indel.save(path)


